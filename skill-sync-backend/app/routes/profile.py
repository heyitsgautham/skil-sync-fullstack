"""
User Profile Routes - Profile management for students, companies, and admins
Feature 7: User Profile Pages
"""

from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel, EmailStr
from datetime import datetime

from app.database.connection import get_db
from app.models import User, UserRole, Internship, Application, Resume
from app.models.student_internship_match import StudentInternshipMatch
from app.utils.security import get_current_user
from app.services.email_service import email_service
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/profile", tags=["Profile"])


# Pydantic schemas
class StudentProfileResponse(BaseModel):
    id: int
    email: str
    full_name: str
    phone: Optional[str] = None
    linkedin_url: Optional[str] = None
    github_url: Optional[str] = None
    skills: Optional[list] = None
    total_experience_years: Optional[float] = None
    
    class Config:
        from_attributes = True


class CompanyProfileResponse(BaseModel):
    id: int
    email: str
    full_name: str  # Company name
    hr_contact_name: Optional[str] = None
    mailing_email: Optional[str] = None
    phone: Optional[str] = None
    phone_visible: bool = False
    
    class Config:
        from_attributes = True


class StudentProfileUpdate(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    linkedin_url: Optional[str] = None
    github_url: Optional[str] = None


class CompanyProfileUpdate(BaseModel):
    full_name: Optional[str] = None  # Company name
    hr_contact_name: Optional[str] = None
    email: Optional[str] = None
    mailing_email: Optional[str] = None
    phone: Optional[str] = None
    phone_visible: Optional[bool] = None


class JobEmailStats(BaseModel):
    internship_id: int
    internship_title: str
    total_applicants: int  # Actually total candidates (all matches)
    great_matches: int  # 80%+
    good_matches: int  # 60-79%
    bad_matches: int  # <60%
    tailored_resume_count: int


@router.get("/me", response_model=dict)
def get_my_profile(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get current user's profile
    
    Returns profile data based on user role (student/company/admin)
    """
    if current_user.role == UserRole.student:
        return {
            "id": current_user.id,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "phone": current_user.phone,
            "linkedin_url": current_user.linkedin_url,
            "github_url": current_user.github_url,
            "skills": current_user.skills,
            "total_experience_years": current_user.total_experience_years,
            "role": current_user.role.value
        }
    elif current_user.role == UserRole.company:
        return {
            "id": current_user.id,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "hr_contact_name": current_user.hr_contact_name,
            "mailing_email": current_user.mailing_email,
            "phone": current_user.phone,
            "phone_visible": current_user.phone_visible,
            "role": current_user.role.value
        }
    else:  # admin
        return {
            "id": current_user.id,
            "email": current_user.email,
            "full_name": current_user.full_name,
            "role": current_user.role.value
        }


@router.put("/me", response_model=dict)
def update_my_profile(
    profile_data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Update current user's profile
    
    Students can update: full_name, phone, linkedin_url, github_url
    Companies can update: full_name, hr_contact_name, email, phone, phone_visible
    """
    try:
        if current_user.role == UserRole.student:
            # Student profile update
            if 'full_name' in profile_data and profile_data['full_name']:
                current_user.full_name = profile_data['full_name']
            if 'phone' in profile_data:
                current_user.phone = profile_data['phone']
            if 'linkedin_url' in profile_data:
                current_user.linkedin_url = profile_data['linkedin_url']
            if 'github_url' in profile_data:
                current_user.github_url = profile_data['github_url']
                
        elif current_user.role == UserRole.company:
            # Company profile update
            if 'full_name' in profile_data and profile_data['full_name']:
                current_user.full_name = profile_data['full_name']
            if 'hr_contact_name' in profile_data:
                current_user.hr_contact_name = profile_data['hr_contact_name']
            if 'email' in profile_data and profile_data['email']:
                # Check if email already exists
                existing_user = db.query(User).filter(
                    User.email == profile_data['email'],
                    User.id != current_user.id
                ).first()
                if existing_user:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Email already in use"
                    )
                current_user.email = profile_data['email']
            if 'mailing_email' in profile_data:
                current_user.mailing_email = profile_data['mailing_email']
            if 'phone' in profile_data:
                current_user.phone = profile_data['phone']
            if 'phone_visible' in profile_data:
                current_user.phone_visible = profile_data['phone_visible']
        
        db.commit()
        db.refresh(current_user)
        
        # Return updated profile
        return get_my_profile(db=db, current_user=current_user)
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update profile: {str(e)}"
        )


@router.get("/job-email-stats")
def get_job_email_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get email statistics for all company's jobs
    Used for the "Send Email" button dropdown
    
    Only accessible by companies
    """
    if current_user.role != UserRole.company:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only companies can access job email statistics"
        )
    
    # Get all internships for this company
    internships = db.query(Internship).filter(
        Internship.company_id == current_user.id,
        Internship.is_active == 1
    ).all()
    
    stats_list = []
    total_stats = {
        "total_applicants": 0,
        "great_matches": 0,
        "good_matches": 0,
        "bad_matches": 0,
        "tailored_resume_count": 0
    }
    
    for internship in internships:
        # Get ALL matched candidates for this internship (not just applicants)
        all_matches = db.query(StudentInternshipMatch).filter(
            StudentInternshipMatch.internship_id == internship.id
        ).all()
        
        # Get applications to track tailored resumes
        applications = db.query(Application).filter(
            Application.internship_id == internship.id
        ).all()
        
        great_matches = 0
        good_matches = 0
        bad_matches = 0
        tailored_count = 0
        
        # Count match quality for ALL candidates
        for match in all_matches:
            score = match.base_similarity_score or 0
            
            if score >= 80:
                great_matches += 1
            elif score >= 60:
                good_matches += 1
            else:
                bad_matches += 1
        
        # Count tailored resumes from applications
        for app in applications:
            if app.used_tailored_resume == 1:
                tailored_count += 1
        
        job_stats = {
            "internship_id": internship.id,
            "internship_title": internship.title,
            "total_applicants": len(all_matches),  # Changed to ALL candidates
            "great_matches": great_matches,
            "good_matches": good_matches,
            "bad_matches": bad_matches,
            "tailored_resume_count": tailored_count
        }
        
        stats_list.append(job_stats)
        
        # Add to totals
        total_stats["total_applicants"] += len(all_matches)
        total_stats["great_matches"] += great_matches
        total_stats["good_matches"] += good_matches
        total_stats["bad_matches"] += bad_matches
        total_stats["tailored_resume_count"] += tailored_count
    
    return {
        "jobs": stats_list,
        "all_jobs_stats": total_stats
    }


@router.post("/send-job-email/{internship_id}")
def send_job_email(
    internship_id: int,
    filters: list[str] = Query(default=["great", "good", "other"], description="Match quality filters: great (≥80%), good (60-79%), other (<60%)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Send email with CSV, Excel, and summary for a specific job
    
    - **internship_id**: ID of the internship (use 0 for all jobs)
    - **filters**: List of match quality filters (great=≥80%, good=60-79%, other=<60%)
    
    Only accessible by companies for their own jobs
    """
    if current_user.role != UserRole.company:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only companies can send job emails"
        )
    
    # Validate filters
    valid_filters = {"great", "good", "other"}
    filters = [f for f in filters if f in valid_filters]
    if not filters:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one valid filter must be provided (great, good, or other)"
        )
    
    # Get stats
    stats_response = get_job_email_stats(db=db, current_user=current_user)
    
    if internship_id == 0:
        # Send email for all jobs
        job_stats = stats_response["all_jobs_stats"]
        internship_title = "All Jobs"
        
        # Get ALL matched candidates for all company internships
        matches_query = db.query(StudentInternshipMatch).join(Internship).filter(
            Internship.company_id == current_user.id,
            Internship.is_active == 1
        )
    else:
        # Verify internship belongs to this company
        internship = db.query(Internship).filter(
            Internship.id == internship_id,
            Internship.company_id == current_user.id
        ).first()
        
        if not internship:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Internship not found or you don't have access"
            )
        
        # Find job stats
        job_stats = next(
            (job for job in stats_response["jobs"] if job["internship_id"] == internship_id),
            None
        )
        
        if not job_stats:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No applications found for this internship"
            )
        
        internship_title = internship.title
        
        # Get ALL matched candidates for this internship
        matches_query = db.query(StudentInternshipMatch).filter(
            StudentInternshipMatch.internship_id == internship_id
        )
    
    all_matches = matches_query.all()
    
    if not all_matches:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No candidates found"
        )
    
    # Filter matches based on match quality filters
    filtered_matches = []
    for match in all_matches:
        score = match.base_similarity_score or 0
        
        if "great" in filters and score >= 80:
            filtered_matches.append(match)
        elif "good" in filters and 60 <= score < 80:
            filtered_matches.append(match)
        elif "other" in filters and score < 60:
            filtered_matches.append(match)
    
    if not filtered_matches:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No candidates found matching the selected filters: {', '.join(filters)}"
        )
    
    # Generate CSV content from filtered matched candidates
    csv_content = generate_csv_export_from_matches(filtered_matches, db, internship)
    
    # Update job_stats to reflect filtered counts
    filtered_stats = {
        "total_applicants": len(filtered_matches),
        "great_matches": sum(1 for m in filtered_matches if (m.base_similarity_score or 0) >= 80),
        "good_matches": sum(1 for m in filtered_matches if 60 <= (m.base_similarity_score or 0) < 80),
        "bad_matches": sum(1 for m in filtered_matches if (m.base_similarity_score or 0) < 60),
        "tailored_resume_count": job_stats.get("tailored_resume_count", 0)
    }
    
    # Generate email HTML with filtered data
    html_content = generate_job_email_html(
        company_name=current_user.full_name,
        internship_title=internship_title,
        job_stats=filtered_stats,
        matches=filtered_matches,
        db=db,
        applied_filters=filters
    )
    
    text_content = generate_job_email_text(
        company_name=current_user.full_name,
        internship_title=internship_title,
        job_stats=filtered_stats
    )
    
    # Generate Excel content from filtered matches
    excel_content = generate_excel_export_from_matches(filtered_matches, db, internship)
    
    # Send email with CSV and Excel attachments
    filter_label = ", ".join([f.upper() for f in filters])
    subject = f"SkillSync Job Report: {internship_title} - {len(filtered_matches)} Candidates ({filter_label})"
    
    # Use mailing_email if set, otherwise fallback to regular email
    recipient_email = current_user.mailing_email or current_user.email
    
    # Create filenames with matching timestamp
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"SkillSync_Candidates_{internship_title.replace(' ', '_')}_{timestamp}.csv"
    xlsx_filename = f"SkillSync_Candidates_{internship_title.replace(' ', '_')}_{timestamp}.xlsx"
    
    # Convert CSV string to bytes
    csv_bytes = csv_content.encode('utf-8')
    
    # Prepare attachments list (both CSV and Excel)
    attachments = [
        (csv_filename, csv_bytes, "text/csv"),
        (xlsx_filename, excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    ]
    
    email_sent = email_service.send_email(
        to_email=recipient_email,
        subject=subject,
        html_content=html_content,
        text_content=text_content,
        attachments=attachments
    )
    
    if not email_sent:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to send email"
        )
    
    return {
        "success": True,
        "message": f"Email sent successfully to {recipient_email}",
        "internship_title": internship_title,
        "total_applicants": len(filtered_matches),
        "applied_filters": filters,
        "csv_preview": csv_content[:500] + "..." if len(csv_content) > 500 else csv_content
    }


def generate_csv_export_from_matches(matches, db, internship):
    """Generate CSV content from matched candidates - matches intelligent_filtering export format"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header - same as intelligent_filtering export
    writer.writerow([
        'Candidate Name', 'Email', 'Phone', 'Match Score (%)',
        'Top Matching Skills', 'Experience (Years)', 'Education Level',
        'Application Date', 'Application Status', 'Key Strengths',
        'Semantic Match (%)', 'Skills Match (%)', 'Experience Match (%)'
    ])
    
    # Sort matches by base_similarity_score descending (highest first)
    sorted_matches = sorted(matches, key=lambda m: m.base_similarity_score or 0, reverse=True)
    
    # Write data
    for match in sorted_matches:
        student = db.query(User).filter(User.id == match.student_id).first()
        if not student:
            continue
        
        # Get the active resume for parsed data
        resume = db.query(Resume).filter(
            Resume.student_id == match.student_id,
            Resume.is_active == 1
        ).first()
        
        # Handle case where resume or parsed_data might be None
        parsed_data = {}
        if resume and resume.parsed_data:
            parsed_data = resume.parsed_data
        personal_info = parsed_data.get('personal_info', {})
        
        # Check if this candidate has applied
        application = db.query(Application).filter(
            Application.student_id == match.student_id,
            Application.internship_id == match.internship_id
        ).first()
        has_applied = 'Yes' if application else 'No'
        
        score = match.base_similarity_score or 0
        
        # Get skills from parsed data and match with required skills
        candidate_skills = parsed_data.get('all_skills', [])
        required_skills = internship.required_skills or []
        matched_skills = [s for s in candidate_skills if s.lower() in [rs.lower() for rs in required_skills]]
        top_skills = ', '.join(matched_skills[:5]) if matched_skills else 'N/A'
        
        # Get education level
        education = parsed_data.get('education', [])
        education_level_str = education[0].get('degree', 'N/A') if education else 'N/A'
        
        # Get key strengths (from projects and certifications)
        projects = parsed_data.get('projects', [])
        certifications = parsed_data.get('certifications', [])
        key_strengths = []
        if projects:
            key_strengths.append(f"{len(projects)} projects")
        if certifications:
            key_strengths.append(f"{len(certifications)} certifications")
        if matched_skills:
            key_strengths.append(f"{len(matched_skills)}/{len(required_skills)} required skills")
        key_strengths_str = ', '.join(key_strengths) if key_strengths else 'Basic qualification'
        
        # Get phone from parsed data
        phone_number = personal_info.get('phone', 'N/A')
        
        # Format phone number for CSV (use formula to force text)
        if phone_number != 'N/A' and phone_number:
            # Ensure phone has proper format with country code
            if not phone_number.startswith('+'):
                phone_number = f"+{phone_number}"
            # Format with space after country code if needed
            if phone_number.startswith('+91-'):
                phone_number = phone_number.replace('+91-', '+91 ')
            # Use Excel formula to force text interpretation
            phone_number = f'="{phone_number}"'
        
        # Get experience from parsed data
        experience_years = parsed_data.get('total_experience_years', 0)
        
        # Get application date and status
        app_date = str(application.created_at) if application else 'Not Applied'
        app_status = application.status if application else 'Not Applied'
        
        writer.writerow([
            student.full_name or 'N/A',
            student.email,
            phone_number,
            f"{score:.2f}",
            top_skills,
            experience_years,
            education_level_str,
            app_date,
            app_status,
            key_strengths_str,
            round(match.semantic_similarity, 2) if match.semantic_similarity is not None else None,
            round(match.skills_match_score, 2) if match.skills_match_score is not None else None,
            round(match.experience_match_score, 2) if match.experience_match_score is not None else None
        ])
    
    return output.getvalue()


def generate_csv_export(applications, db):
    """Generate CSV content from applications (legacy - for backwards compatibility)"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Candidate Name', 'Email', 'Phone', 'Match Score (%)',
        'Tailored Resume', 'Top Skills', 'Experience (Years)',
        'Application Date', 'Status'
    ])
    
    # Write data
    for app in applications:
        student = db.query(User).filter(User.id == app.student_id).first()
        if not student:
            continue
        
        score = app.application_similarity_score or app.match_score or 0
        skills_str = ', '.join(student.skills[:5]) if student.skills else 'N/A'
        
        writer.writerow([
            student.full_name,
            student.email,
            student.phone or 'N/A',
            score,
            'Yes' if app.used_tailored_resume == 1 else 'No',
            skills_str,
            student.total_experience_years or 0,
            app.created_at.strftime('%Y-%m-%d %H:%M'),
            app.status
        ])
    
    return output.getvalue()


def generate_excel_export_from_matches(matches, db, internship):
    """Generate Excel (XLSX) content from matched candidates - matches intelligent_filtering export format"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Rankings"
    
    # Define styles
    from openpyxl.styles import Border, Side
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title row (row 1)
    ws.merge_cells('A1:M1')
    ws['A1'] = f"Candidate Rankings - {internship.title}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add summary row (row 2)
    ws.merge_cells('A2:M2')
    ws['A2'] = f"Exported on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Candidates: {len(matches)}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Define headers - same as intelligent_filtering export
    headers = [
        'Candidate Name', 'Email', 'Phone', 'Match Score (%)',
        'Top Matching Skills', 'Experience (Years)', 'Education Level',
        'Application Date', 'Application Status', 'Key Strengths',
        'Semantic Match (%)', 'Skills Match (%)', 'Experience Match (%)'
    ]
    
    # Write headers with styling (row 4)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Sort matches by base_similarity_score descending (highest first)
    sorted_matches = sorted(matches, key=lambda m: m.base_similarity_score or 0, reverse=True)
    
    # Write data rows
    row_num = 5
    for match in sorted_matches:
        student = db.query(User).filter(User.id == match.student_id).first()
        if not student:
            continue
        
        # Get the active resume for parsed data
        resume = db.query(Resume).filter(
            Resume.student_id == match.student_id,
            Resume.is_active == 1
        ).first()
        
        # Handle case where resume or parsed_data might be None
        parsed_data = {}
        if resume and resume.parsed_data:
            parsed_data = resume.parsed_data
        personal_info = parsed_data.get('personal_info', {})
        
        # Check if this candidate has applied
        application = db.query(Application).filter(
            Application.student_id == match.student_id,
            Application.internship_id == match.internship_id
        ).first()
        has_applied = 'Yes' if application else 'No'
        
        score = match.base_similarity_score or 0
        
        # Get skills from parsed data and match with required skills
        candidate_skills = parsed_data.get('all_skills', [])
        required_skills = internship.required_skills or []
        matched_skills = [s for s in candidate_skills if s.lower() in [rs.lower() for rs in required_skills]]
        top_skills = ', '.join(matched_skills[:5]) if matched_skills else 'N/A'
        
        # Get education level
        education = parsed_data.get('education', [])
        education_level_str = education[0].get('degree', 'N/A') if education else 'N/A'
        
        # Get key strengths (from projects and certifications)
        projects = parsed_data.get('projects', [])
        certifications = parsed_data.get('certifications', [])
        key_strengths = []
        if projects:
            key_strengths.append(f"{len(projects)} projects")
        if certifications:
            key_strengths.append(f"{len(certifications)} certifications")
        if matched_skills:
            key_strengths.append(f"{len(matched_skills)}/{len(required_skills)} required skills")
        key_strengths_str = ', '.join(key_strengths) if key_strengths else 'Basic qualification'
        
        # Get phone from parsed data
        phone_number = personal_info.get('phone', 'N/A')
        
        # Format phone number for Excel
        if phone_number != 'N/A' and phone_number:
            # Ensure phone has proper format with country code
            if not phone_number.startswith('+'):
                phone_number = f"+{phone_number}"
            # Format with space after country code if needed
            if phone_number.startswith('+91-'):
                phone_number = phone_number.replace('+91-', '+91 ')
        
        # Get experience from parsed data
        experience_years = parsed_data.get('total_experience_years', 0)
        
        # Get application date and status
        app_date = str(application.created_at) if application else 'Not Applied'
        app_status = application.status if application else 'Not Applied'
        
        # Write row data
        ws.cell(row=row_num, column=1, value=student.full_name or 'N/A')
        ws.cell(row=row_num, column=2, value=student.email)
        
        # Format phone number and set as text
        phone_cell = ws.cell(row=row_num, column=3, value=phone_number)
        phone_cell.number_format = '@'  # @ means text format in Excel to prevent formula interpretation
        
        ws.cell(row=row_num, column=4, value=round(score, 2))
        ws.cell(row=row_num, column=5, value=top_skills)
        ws.cell(row=row_num, column=6, value=experience_years)
        ws.cell(row=row_num, column=7, value=education_level_str)
        ws.cell(row=row_num, column=8, value=app_date)
        ws.cell(row=row_num, column=9, value=app_status)
        ws.cell(row=row_num, column=10, value=key_strengths_str)
        ws.cell(row=row_num, column=11, value=round(match.semantic_similarity, 2) if match.semantic_similarity is not None else None)
        ws.cell(row=row_num, column=12, value=round(match.skills_match_score, 2) if match.skills_match_score is not None else None)
        ws.cell(row=row_num, column=13, value=round(match.experience_match_score, 2) if match.experience_match_score is not None else None)
        
        # Color code only the Match Score cell
        score_cell = ws.cell(row=row_num, column=4)
        if score >= 80:
            score_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
        elif score >= 60:
            score_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # Light orange
        else:
            score_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Light red
        
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_job_email_html(company_name, internship_title, job_stats, matches, db, applied_filters=None):
    """Generate HTML email content for job report"""
    
    # Generate filter info HTML if filters were applied
    filter_info = ""
    if applied_filters and len(applied_filters) < 3:  # If not all filters selected
        filter_labels = {
            "great": "Great Matches (≥80%)",
            "good": "Good Matches (60-79%)",
            "other": "Other Matches (<60%)"
        }
        selected_filters = [filter_labels.get(f, f.upper()) for f in applied_filters]
        filter_info = f"""
            <div style="background-color: #e3f2fd; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <strong>📋 Filtered Categories:</strong> {', '.join(selected_filters)}
                <br/>
                <em style="color: #666; font-size: 13px;">This report includes only candidates from the selected match quality categories.</em>
            </div>
        """
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 800px;
                margin: 0 auto;
                padding: 20px;
            }}
            .header {{
                background-color: #1976d2;
                color: white;
                padding: 20px;
                text-align: center;
                border-radius: 8px 8px 0 0;
            }}
            .content {{
                background-color: #f5f5f5;
                padding: 30px;
                border-radius: 0 0 8px 8px;
            }}
            .stats-grid {{
                display: grid;
                grid-template-columns: repeat(2, 1fr);
                gap: 15px;
                margin: 20px 0;
            }}
            .stat-card {{
                background-color: white;
                padding: 15px;
                border-radius: 8px;
                border-left: 4px solid #1976d2;
            }}
            .stat-number {{
                font-size: 32px;
                font-weight: bold;
                color: #1976d2;
            }}
            .stat-label {{
                color: #666;
                font-size: 14px;
            }}
            .great {{ border-left-color: #4caf50; }}
            .good {{ border-left-color: #ff9800; }}
            .bad {{ border-left-color: #f44336; }}
            .tailored {{ border-left-color: #9c27b0; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>📊 Job Candidate Report</h1>
            <h2>{internship_title}</h2>
        </div>
        <div class="content">
            <h3>Hello, {company_name}!</h3>
            <p>Here's your comprehensive report for <strong>{internship_title}</strong>.</p>
            
            {filter_info}
            
            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-number">{job_stats['total_applicants']}</div>
                    <div class="stat-label">Total Candidates</div>
                </div>
                <div class="stat-card great">
                    <div class="stat-number">{job_stats['great_matches']}</div>
                    <div class="stat-label">Great Matches (≥80%)</div>
                </div>
                <div class="stat-card good">
                    <div class="stat-number">{job_stats['good_matches']}</div>
                    <div class="stat-label">Good Matches (60-79%)</div>
                </div>
                <div class="stat-card bad">
                    <div class="stat-number">{job_stats['bad_matches']}</div>
                    <div class="stat-label">Other Matches (<60%)</div>
                </div>
                <div class="stat-card tailored">
                    <div class="stat-number">{job_stats['tailored_resume_count']}</div>
                    <div class="stat-label">Tailored Resumes</div>
                </div>
            </div>
            
            <p><strong>Note:</strong> CSV and Excel exports with detailed candidate information are attached to this email.</p>
            
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #ddd;">
            
            <p style="font-size: 12px; color: #666; text-align: center;">
                This is an automated report from SkillSync<br>
                Feature 7: User Profile Pages - Job Email Reports<br>
                Date: {datetime.utcnow().strftime('%B %d, %Y')}
            </p>
        </div>
    </body>
    </html>
    """
    return html


def generate_job_email_text(company_name, internship_title, job_stats):
    """Generate plain text email content for job report"""
    text = f"""
SKILLSYNC - JOB CANDIDATE REPORT
{internship_title}
{'='*60}

Hello, {company_name}!

Here's your comprehensive report for {internship_title}.

STATISTICS:
-----------
Total Candidates:        {job_stats['total_applicants']}
Great Matches (80%+):    {job_stats['great_matches']}
Good Matches (60-79%):   {job_stats['good_matches']}
Low Matches (<60%):      {job_stats['bad_matches']}
Tailored Resumes:        {job_stats['tailored_resume_count']}

Note: CSV export with detailed candidate information is attached to this email.

{'='*60}

This is an automated report from SkillSync
Feature 7: User Profile Pages - Job Email Reports
Date: {datetime.utcnow().strftime('%B %d, %Y')}
    """
    return text
